#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MQTT话题监控器 - Excel实时记录系统
每个话题对应一个Sheet页，实时监控数据变化并记录到Excel
"""

import json
import time
import datetime
import threading
import signal
from typing import Dict, Any, Optional
import paho.mqtt.client as mqtt
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os
import sys

class MQTTExcelMonitor:
    def __init__(self, excel_file="mqtt_monitor.xlsx", mqtt_host="hellorobotaxi.cn", mqtt_port=11883):
        self.excel_file = excel_file
        self.mqtt_host = mqtt_host
        self.mqtt_port = mqtt_port
        
        # 存储每个话题的最新数据，用于变化检测
        self.last_data = {}
        
        # 存储每个话题对应的sheet名称
        self.topic_sheets = {}
        
        # MQTT客户端
        self.mqtt_client = mqtt.Client()
        self.mqtt_client.on_connect = self.on_connect
        self.mqtt_client.on_message = self.on_message
        
        # Excel工作簿
        self.workbook = None
        self.excel_lock = threading.Lock()
        
        # 运行状态控制
        self.running = True
        
        # 初始化Excel文件
        self.init_excel()
        
        print(f"[SUCCESS] MQTT Excel监控器初始化完成")
        print(f"Excel文件: {os.path.abspath(self.excel_file)}")
        print(f"MQTT服务器: {self.mqtt_host}:{self.mqtt_port}")

    def topic_to_sheet_name(self, topic: str) -> str:
        """将MQTT话题名转换为有效的Excel sheet名称"""
        # 替换特殊字符
        sheet_name = topic.replace("/", "_").replace("+", "wildcard")
        # 移除开头的下划线
        if sheet_name.startswith("_"):
            sheet_name = sheet_name[1:]
        # 限制长度（Excel sheet名称最大31字符）
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:28] + "..."
        return sheet_name

    def init_excel(self):
        """初始化Excel文件"""
        try:
            if os.path.exists(self.excel_file):
                self.workbook = openpyxl.load_workbook(self.excel_file)
                print(f"[SUCCESS] 加载现有Excel文件: {self.excel_file}")
            else:
                self.workbook = Workbook()
                # 删除默认的Sheet
                self.workbook.remove(self.workbook.active)
                print(f"[SUCCESS] 创建新Excel文件: {self.excel_file}")
                
        except Exception as e:
            print(f"[ERROR] Excel文件初始化失败: {e}")
            self.workbook = Workbook()

    def get_or_create_sheet(self, topic: str):
        """获取或创建指定话题的sheet页"""
        sheet_name = self.topic_to_sheet_name(topic)
        
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        
        # 创建新的sheet页
        sheet = self.workbook.create_sheet(title=sheet_name)
        
        # 设置表头
        headers = ["字段名", "当前值", "更新时间", "备注"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # 调整列宽
        sheet.column_dimensions['A'].width = 25  # 字段名
        sheet.column_dimensions['B'].width = 20  # 当前值
        sheet.column_dimensions['C'].width = 20  # 更新时间
        sheet.column_dimensions['D'].width = 30  # 备注
        
        print(f"[SUCCESS] 创建新sheet页: {sheet_name} (话题: {topic})")
        return sheet

    def detect_changes(self, topic: str, new_data: Dict[str, Any]) -> list:
        """检测数据变化"""
        changes = []
        
        if topic not in self.last_data:
            # 首次接收到该话题的数据
            for field, value in new_data.items():
                changes.append({
                    'field': field,
                    'old_value': None,
                    'new_value': value,
                    'change_type': '首次接收'
                })
        else:
            # 比较与上次数据的差异
            old_data = self.last_data[topic]
            
            # 检查已有字段的变化
            for field, new_value in new_data.items():
                old_value = old_data.get(field)
                if old_value != new_value:
                    change_type = '数值变化' if old_value is not None else '新增字段'
                    changes.append({
                        'field': field,
                        'old_value': old_value,
                        'new_value': new_value,
                        'change_type': change_type
                    })
            
            # 检查删除的字段
            for field in old_data:
                if field not in new_data:
                    changes.append({
                        'field': field,
                        'old_value': old_data[field],
                        'new_value': None,
                        'change_type': '字段删除'
                    })
        
        return changes

    def update_field_in_excel(self, topic: str, field_name: str, new_value: Any):
        """更新Excel中指定字段的值和时间戳"""
        with self.excel_lock:
            try:
                sheet = self.get_or_create_sheet(topic)
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # 查找是否已存在该字段
                field_row = None
                for row in range(2, sheet.max_row + 1):
                    if sheet.cell(row=row, column=1).value == field_name:
                        field_row = row
                        break
                
                if field_row is None:
                    # 新字段，添加到最后
                    field_row = sheet.max_row + 1
                    sheet.cell(row=field_row, column=1, value=field_name)  # 字段名
                    sheet.cell(row=field_row, column=4, value=f"话题: {topic}")  # 备注
                
                # 更新当前值和时间戳
                sheet.cell(row=field_row, column=2, value=str(new_value))  # 当前值
                sheet.cell(row=field_row, column=3, value=timestamp)       # 更新时间
                
                # 保存文件
                self.workbook.save(self.excel_file)
                
            except Exception as e:
                print(f"[ERROR] 更新Excel失败: {e}")

    def write_changes_to_excel(self, topic: str, changes: list):
        """将变化写入Excel"""
        if not changes:
            return
        
        for change in changes:
            self.update_field_in_excel(topic, change['field'], change['new_value'])
        
        print(f"[UPDATE] {topic} - 更新{len(changes)}个字段到Excel")

    def on_connect(self, client, userdata, flags, rc):
        """MQTT连接回调"""
        if rc == 0:
            print(f"[SUCCESS] 连接MQTT服务器成功: {self.mqtt_host}:{self.mqtt_port}")
            
            # 订阅关键话题
            topics_to_monitor = [
                "/handshake/request",
                "/handshake/response", 
                "/vehicle/vehicle_status",
                "/vehicle/control_cmd",
                "/vehicle/vehicle_status/+",  # VID话题
                "/handshake/request/+",       # VID话题
                "tsp/command/+",
                "cloud/noa/+",
                "zhan/send/+",
                "/handshake/request/noa"
            ]
            
            for topic in topics_to_monitor:
                client.subscribe(topic)
                print(f"[SUCCESS] 订阅话题: {topic}")
                
        else:
            print(f"[ERROR] MQTT连接失败，错误码: {rc}")

    def on_message(self, client, userdata, msg):
        """MQTT消息回调"""
        try:
            topic = msg.topic
            payload = msg.payload.decode('utf-8')
            
            # 解析JSON数据
            try:
                data = json.loads(payload)
            except json.JSONDecodeError:
                # 如果不是JSON格式，作为字符串处理
                data = {"raw_message": payload}
            
            # 检测变化
            changes = self.detect_changes(topic, data)
            
            if changes:
                # 有变化时才记录
                self.write_changes_to_excel(topic, changes)
                
                # 简要打印变化信息
                change_summary = f"{topic}: {len(changes)}个字段变化"
                if len(changes) <= 3:
                    field_names = [change['field'] for change in changes]
                    change_summary += f" ({', '.join(field_names)})"
                print(f"[CHANGE] {change_summary}")
            
            # 更新最新数据
            self.last_data[topic] = data.copy()
            
        except Exception as e:
            print(f"[ERROR] 处理消息失败: {e}")

    def start_monitoring(self):
        """开始监控"""
        try:
            self.mqtt_client.connect(self.mqtt_host, self.mqtt_port, 60)
            
            # 使用loop_start以便可以响应信号
            self.mqtt_client.loop_start()
            
            # 主循环，等待停止信号
            while self.running:
                time.sleep(0.1)
                
        except Exception as e:
            print(f"[ERROR] MQTT监控启动失败: {e}")
        finally:
            self.stop_monitoring()

    def stop_monitoring(self):
        """停止监控"""
        self.running = False
        self.mqtt_client.loop_stop()
        self.mqtt_client.disconnect()
        if self.workbook:
            with self.excel_lock:
                self.workbook.save(self.excel_file)
        print(f"[SUCCESS] 监控已停止，Excel文件已保存: {os.path.abspath(self.excel_file)}")

    def signal_handler(self, signum, frame):
        """信号处理器"""
        print(f"\n[INFO] 收到停止信号 {signum}，正在优雅退出...")
        self.stop_monitoring()
        sys.exit(0)

def main():
    print("=== MQTT话题监控器 - Excel实时记录系统 ===")
    print("每个话题对应一个Sheet页，实时监控数据变化")
    print("按 Ctrl+C 停止监控")
    print("")
    
    # 创建监控器
    monitor = MQTTExcelMonitor()
    
    # 注册信号处理器
    signal.signal(signal.SIGINT, monitor.signal_handler)
    signal.signal(signal.SIGTERM, monitor.signal_handler)
    
    try:
        # 开始监控
        monitor.start_monitoring()
    except KeyboardInterrupt:
        print("\n收到停止信号...")
        monitor.stop_monitoring()
    except Exception as e:
        print(f"[ERROR] 程序异常: {e}")
        monitor.stop_monitoring()

if __name__ == "__main__":
    main()
